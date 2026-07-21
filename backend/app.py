"""
Khorshid (خورشید) — Home Appliance Distribution & Repair Co.
Backend API — Flask

Responsibilities
- Reads live product data (name / price / stock) from an Excel file using openpyxl
  and serves it as JSON at GET /api/products. The file is re-read on every request
  (cheap for a catalog this size) so any edit saved in Excel is reflected on the
  very next poll — the frontend polls every 10 seconds. This is deliberate and
  NOT part of the SQLite migration below.
- Auth (signup / login / logout / forgot-password) backed by SQLite (db.py),
  password hashed with werkzeug's generator. A small brute-force guard locks an
  account for 15 minutes after 5 bad password attempts.
- Orders, invoices, coupons and repair requests all live in SQLite now (see
  db.py) instead of flat JSON files — same data, real transactions/indexes.
- Admin area (session user with is_admin=1): view/update orders, manage
  coupons, view repair requests, see the accounting summary. A default admin
  account is created on first run — see README for the credentials and how
  to change them.
- Shipping rule: a customer's first order ships free; every order after that
  has a flat shipping fee. Coupons apply a percentage discount on top.

Run:
    pip install -r requirements.txt
    python app.py
Then open http://localhost:5000
"""

import os
import re
import uuid
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path

from flask import Flask, jsonify, request, session, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl

import db

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
FRONTEND_DIR = BASE_DIR.parent / "frontend"

PRODUCTS_XLSX = DATA_DIR / "products.xlsx"
AVATAR_DIR = DATA_DIR / "avatars"
AVATAR_DIR.mkdir(parents=True, exist_ok=True)
ALLOWED_AVATAR_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "gif"}

FREE_SHIPPING_ON_FIRST_ORDER = True
STANDARD_SHIPPING_FEE = 350_000  # تومان

app = Flask(__name__, static_folder=None)
app.secret_key = os.environ.get("KHORSHID_SECRET_KEY", "dev-secret-change-me")
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10MB, covers avatar/product uploads

db.init_db()


# --------------------------------------------------------------------------- #
# auth helpers
# --------------------------------------------------------------------------- #
def current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    return db.get_user_by_id(uid)


def require_admin(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = current_user()
        if not user or not user["is_admin"]:
            return jsonify({"ok": False, "error": "admin_required"}), 403
        return fn(*args, **kwargs)
    return wrapper


# --------------------------------------------------------------------------- #
# products (read live from Excel) — unchanged from before, not part of SQLite
# --------------------------------------------------------------------------- #
def _split_gallery(raw):
    """gallery_images column holds comma/newline/pipe-separated URLs — split
    on any of those and drop blanks, so editors can use whichever separator
    is natural in Excel."""
    if not raw:
        return []
    import re
    parts = re.split(r"[,\n|]+", str(raw))
    return [p.strip() for p in parts if p.strip()]


def _split_specs(raw):
    """specs_fa / specs_en columns hold one 'label: value' pair per line
    (Alt+Enter inside the Excel cell) — parsed into an ordered list of
    {label, value} for the frontend's spec table."""
    if not raw:
        return []
    specs = []
    for line in str(raw).splitlines():
        line = line.strip()
        if not line:
            continue
        if ":" in line:
            label, value = line.split(":", 1)
            specs.append({"label": label.strip(), "value": value.strip()})
        else:
            specs.append({"label": line, "value": ""})
    return specs


def read_products_from_excel():
    if not PRODUCTS_XLSX.exists():
        return []

    wb = openpyxl.load_workbook(PRODUCTS_XLSX, data_only=True, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    products = []
    for row in rows[1:]:
        if row is None or all(v is None for v in row):
            continue
        record = dict(zip(headers, row))
        try:
            main_image = record.get("image") or ""
            gallery = _split_gallery(record.get("gallery_images"))
            # the main image doubles as the first gallery/thumbnail entry
            full_gallery = [main_image] + [g for g in gallery if g != main_image] if main_image else gallery

            products.append({
                "id": int(record.get("id")),
                "name_fa": record.get("name_fa") or "",
                "name_en": record.get("name_en") or "",
                "category_fa": record.get("category_fa") or "",
                "category_en": record.get("category_en") or "",
                "price": int(record.get("price_toman") or 0),
                "stock": int(record.get("stock") or 0),
                "image": main_image,
                "sku": record.get("sku") or "",
                "description_fa": record.get("description_fa") or "",
                "description_en": record.get("description_en") or "",
                "specs_fa": _split_specs(record.get("specs_fa")),
                "specs_en": _split_specs(record.get("specs_en")),
                "gallery": full_gallery,
            })
        except (TypeError, ValueError):
            continue
    wb.close()
    return products


def decrement_stock(items):
    if not PRODUCTS_XLSX.exists() or not items:
        return

    wb = openpyxl.load_workbook(PRODUCTS_XLSX, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    id_col = headers.index("id") + 1
    stock_col = headers.index("stock") + 1

    qty_by_id = {}
    for i in items:
        qty_by_id[int(i["id"])] = qty_by_id.get(int(i["id"]), 0) + int(i.get("qty", 1))

    for row in ws.iter_rows(min_row=2):
        cell_id = row[id_col - 1].value
        if cell_id in qty_by_id:
            current = row[stock_col - 1].value or 0
            row[stock_col - 1].value = max(0, int(current) - qty_by_id[cell_id])

    wb.save(PRODUCTS_XLSX)
    wb.close()


@app.route("/api/products", methods=["GET"])
def api_products():
    try:
        products = read_products_from_excel()
        mtime = PRODUCTS_XLSX.stat().st_mtime if PRODUCTS_XLSX.exists() else 0
        return jsonify({
            "ok": True,
            "updated_at": mtime,
            "count": len(products),
            "products": products,
        })
    except Exception as exc:
        msg = str(exc)
        if isinstance(exc, PermissionError):
            msg = (
                "products.xlsx is locked by another program — close it in Excel "
                "(or any app that has it open) and try again."
            )
        return jsonify({"ok": False, "error": msg, "products": []}), 200


@app.route("/api/admin/upload-products", methods=["POST"])
@require_admin
def api_upload_products():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "no file provided"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "error": "file must be .xlsx"}), 400
    f.save(PRODUCTS_XLSX)
    return jsonify({"ok": True, "message": "products.xlsx updated"})


@app.route("/api/products/<int:product_id>", methods=["GET"])
def api_product_detail(product_id):
    try:
        products = read_products_from_excel()
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 200

    product = next((p for p in products if p["id"] == product_id), None)
    if not product:
        return jsonify({"ok": False, "error": "product_not_found"}), 404

    related = [
        p for p in products
        if p["id"] != product_id and p["category_fa"] == product["category_fa"]
    ][:8]
    if not related:
        # no other product shares this category (common with a small catalog)
        # — fall back to other products so the "similar products" section
        # isn't just empty on every page
        related = [p for p in products if p["id"] != product_id][:8]

    return jsonify({"ok": True, "product": product, "related": related})


# --------------------------------------------------------------------------- #
# product reviews (stored server-side in SQLite — never in the browser)
# --------------------------------------------------------------------------- #
@app.route("/api/products/<int:product_id>/reviews", methods=["GET"])
def api_get_reviews(product_id):
    reviews = db.get_reviews_for_product(product_id)
    summary = db.get_review_summary(product_id)
    my_review = None
    user = current_user()
    if user:
        my_review = db.get_user_review_for_product(product_id, user["id"])
    return jsonify({"ok": True, "reviews": reviews, "summary": summary, "my_review": my_review})


@app.route("/api/products/<int:product_id>/reviews", methods=["POST"])
def api_post_review(product_id):
    user = current_user()
    if not user:
        return jsonify({"ok": False, "error": "login_required"}), 401

    body = request.get_json(force=True, silent=True) or {}
    try:
        rating = int(body.get("rating"))
    except (TypeError, ValueError):
        rating = None
    comment = (body.get("comment") or "").strip()

    if rating is None or not (1 <= rating <= 5):
        return jsonify({"ok": False, "error": "invalid_rating"}), 400
    if len(comment) > 2000:
        return jsonify({"ok": False, "error": "comment_too_long"}), 400

    products = read_products_from_excel()
    if not any(p["id"] == product_id for p in products):
        return jsonify({"ok": False, "error": "product_not_found"}), 404

    db.upsert_review(product_id, user["id"], user["name"], rating, comment)
    reviews = db.get_reviews_for_product(product_id)
    summary = db.get_review_summary(product_id)
    my_review = db.get_user_review_for_product(product_id, user["id"])
    return jsonify({"ok": True, "reviews": reviews, "summary": summary, "my_review": my_review})


# --------------------------------------------------------------------------- #
# auth
# --------------------------------------------------------------------------- #
@app.route("/api/auth/signup", methods=["POST"])
def signup():
    body = request.get_json(force=True, silent=True) or {}
    first_name = (body.get("first_name") or "").strip()
    last_name = (body.get("last_name") or "").strip()
    email = (body.get("email") or "").strip().lower()
    password = body.get("password") or ""

    if not first_name or not last_name or not email or len(password) < 6:
        return jsonify({"ok": False, "error": "invalid_input"}), 400

    if db.get_user_by_email(email):
        return jsonify({"ok": False, "error": "email_exists"}), 409

    user = db.create_user(first_name, last_name, email, generate_password_hash(password))
    session["user_id"] = user["id"]
    return jsonify({"ok": True, "user": db.user_to_dict(user)})


@app.route("/api/auth/login", methods=["POST"])
def login():
    body = request.get_json(force=True, silent=True) or {}
    email = (body.get("email") or "").strip().lower()
    password = body.get("password") or ""

    user = db.get_user_by_email(email)
    if not user:
        return jsonify({"ok": False, "error": "invalid_credentials"}), 401

    if db.is_locked(user):
        return jsonify({"ok": False, "error": "account_locked"}), 423

    if not check_password_hash(user["password_hash"], password):
        db.record_login_attempt(email, success=False)
        return jsonify({"ok": False, "error": "invalid_credentials"}), 401

    db.record_login_attempt(email, success=True)
    session["user_id"] = user["id"]
    return jsonify({"ok": True, "user": db.user_to_dict(user)})


@app.route("/api/auth/logout", methods=["POST"])
def logout():
    session.pop("user_id", None)
    return jsonify({"ok": True})


@app.route("/api/auth/me", methods=["GET"])
def me():
    user = current_user()
    return jsonify({"ok": True, "user": db.user_to_dict(user)})


@app.route("/api/auth/forgot-password", methods=["POST"])
def forgot_password():
    """
    Demo-only flow: there's no email/SMS service wired up, so instead of
    emailing a reset link we hand the token straight back in the response
    for the frontend to show on screen. Swap this for a real email/SMS send
    (see README) before using this in production — never expose reset
    tokens in the API response for a real deployment.
    """
    body = request.get_json(force=True, silent=True) or {}
    email = (body.get("email") or "").strip().lower()
    user = db.get_user_by_email(email)
    if not user:
        # don't reveal whether the email exists
        return jsonify({"ok": True, "demo_note": "email_not_found_but_hidden"})

    token = uuid.uuid4().hex
    expires = (datetime.utcnow() + timedelta(minutes=30)).isoformat()
    db.set_reset_token(email, token, expires)
    return jsonify({"ok": True, "demo_reset_token": token})


@app.route("/api/auth/reset-password", methods=["POST"])
def reset_password():
    body = request.get_json(force=True, silent=True) or {}
    token = body.get("token") or ""
    new_password = body.get("password") or ""
    if len(new_password) < 6:
        return jsonify({"ok": False, "error": "invalid_input"}), 400

    user = db.get_user_by_reset_token(token)
    if not user or not user["reset_token_expires"]:
        return jsonify({"ok": False, "error": "invalid_token"}), 400
    if datetime.fromisoformat(user["reset_token_expires"]) < datetime.utcnow():
        return jsonify({"ok": False, "error": "token_expired"}), 400

    db.reset_password(user["id"], generate_password_hash(new_password))
    return jsonify({"ok": True})


# --------------------------------------------------------------------------- #
# account / profile (name, phone, saved address, password, avatar)
# --------------------------------------------------------------------------- #
@app.route("/api/profile/update", methods=["POST"])
def update_profile_route():
    user = current_user()
    if not user:
        return jsonify({"ok": False, "error": "login_required"}), 401

    body = request.get_json(force=True, silent=True) or {}
    first_name = (body.get("first_name") or "").strip()
    last_name = (body.get("last_name") or "").strip()
    phone = (body.get("phone") or "").strip()
    if not first_name or not last_name:
        return jsonify({"ok": False, "error": "invalid_input"}), 400

    db.update_profile(user["id"], first_name, last_name, phone)
    return jsonify({"ok": True, "user": db.user_to_dict(db.get_user_by_id(user["id"]))})


@app.route("/api/profile/address", methods=["POST"])
def update_address_route():
    user = current_user()
    if not user:
        return jsonify({"ok": False, "error": "login_required"}), 401

    body = request.get_json(force=True, silent=True) or {}
    lat, lng = body.get("lat"), body.get("lng")
    address_text = (body.get("address") or "").strip()
    if lat is None or lng is None:
        return jsonify({"ok": False, "error": "invalid_input"}), 400

    try:
        lat, lng = float(lat), float(lng)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "invalid_input"}), 400

    db.update_address(user["id"], address_text, lat, lng)
    return jsonify({"ok": True, "user": db.user_to_dict(db.get_user_by_id(user["id"]))})


@app.route("/api/profile/password", methods=["POST"])
def change_password_route():
    user = current_user()
    if not user:
        return jsonify({"ok": False, "error": "login_required"}), 401

    body = request.get_json(force=True, silent=True) or {}
    current_password = body.get("current_password") or ""
    new_password = body.get("new_password") or ""
    if len(new_password) < 6:
        return jsonify({"ok": False, "error": "invalid_input"}), 400
    if not check_password_hash(user["password_hash"], current_password):
        return jsonify({"ok": False, "error": "invalid_current_password"}), 401

    db.set_password(user["id"], generate_password_hash(new_password))
    return jsonify({"ok": True})


@app.route("/api/profile/avatar", methods=["POST"])
def upload_avatar_route():
    user = current_user()
    if not user:
        return jsonify({"ok": False, "error": "login_required"}), 401

    if "file" not in request.files or not request.files["file"].filename:
        return jsonify({"ok": False, "error": "no_file"}), 400

    f = request.files["file"]
    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in ALLOWED_AVATAR_EXTENSIONS:
        return jsonify({"ok": False, "error": "invalid_file_type"}), 400

    # drop any older avatar saved under a different extension for this user
    for old_file in AVATAR_DIR.glob(f"{user['id']}.*"):
        old_file.unlink(missing_ok=True)

    filename = f"{user['id']}.{ext}"
    f.save(AVATAR_DIR / filename)
    db.update_avatar(user["id"], filename)
    return jsonify({"ok": True, "user": db.user_to_dict(db.get_user_by_id(user["id"]))})


@app.route("/api/avatars/<path:filename>")
def serve_avatar(filename):
    return send_from_directory(AVATAR_DIR, filename)


# --------------------------------------------------------------------------- #
# coupons
# --------------------------------------------------------------------------- #
@app.route("/api/coupons/validate", methods=["POST"])
def validate_coupon():
    body = request.get_json(force=True, silent=True) or {}
    code = (body.get("code") or "").strip()
    coupon = db.get_coupon(code) if code else None
    if not coupon or not coupon["active"]:
        return jsonify({"ok": False, "error": "invalid_coupon"}), 404
    return jsonify({"ok": True, "code": coupon["code"], "discount_percent": coupon["discount_percent"]})


@app.route("/api/admin/coupons", methods=["GET"])
@require_admin
def admin_list_coupons():
    return jsonify({"ok": True, "coupons": db.get_all_coupons()})


@app.route("/api/admin/coupons", methods=["POST"])
@require_admin
def admin_create_coupon():
    body = request.get_json(force=True, silent=True) or {}
    code = (body.get("code") or "").strip()
    try:
        percent = int(body.get("discount_percent"))
    except (TypeError, ValueError):
        percent = None
    if not code or percent is None or not (0 < percent <= 90):
        return jsonify({"ok": False, "error": "invalid_input"}), 400
    db.create_coupon(code, percent)
    return jsonify({"ok": True})


@app.route("/api/delivery/blocked-dates", methods=["GET"])
def api_blocked_dates():
    """Public — the checkout page uses this to grey out/disable dates the
    customer shouldn't be able to pick (e.g. holidays or fully-booked days)."""
    dates = [row["date"] for row in db.get_blocked_dates()]
    return jsonify({"ok": True, "dates": dates})


@app.route("/api/admin/blocked-dates", methods=["GET"])
@require_admin
def admin_list_blocked_dates():
    return jsonify({"ok": True, "dates": db.get_blocked_dates()})


@app.route("/api/admin/blocked-dates", methods=["POST"])
@require_admin
def admin_block_date():
    body = request.get_json(force=True, silent=True) or {}
    date_str = (body.get("date") or "").strip()
    reason = (body.get("reason") or "").strip() or None
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", date_str):
        return jsonify({"ok": False, "error": "invalid_date"}), 400
    db.block_delivery_date(date_str, reason)
    return jsonify({"ok": True})


@app.route("/api/admin/blocked-dates/<date_str>/unblock", methods=["POST"])
@require_admin
def admin_unblock_date(date_str):
    db.unblock_delivery_date(date_str)
    return jsonify({"ok": True})


@app.route("/api/admin/coupons/<code>/toggle", methods=["POST"])
@require_admin
def admin_toggle_coupon(code):
    coupon = db.get_coupon(code)
    if not coupon:
        return jsonify({"ok": False, "error": "not_found"}), 404
    db.set_coupon_active(code, not coupon["active"])
    return jsonify({"ok": True})


# --------------------------------------------------------------------------- #
# orders / shipping / accounting
# --------------------------------------------------------------------------- #
@app.route("/api/orders", methods=["POST"])
def create_order_route():
    uid = session.get("user_id")
    if not uid:
        return jsonify({"ok": False, "error": "login_required"}), 401

    body = request.get_json(force=True, silent=True) or {}
    items = body.get("items") or []
    delivery = body.get("delivery") or {}
    customer = body.get("customer") or {}
    coupon_code = (body.get("coupon_code") or "").strip()

    if not items:
        return jsonify({"ok": False, "error": "empty_cart"}), 400

    if not str(customer.get("phone") or "").strip():
        return jsonify({"ok": False, "error": "phone_required"}), 400

    slot_date = (delivery.get("slot") or {}).get("date") if isinstance(delivery.get("slot"), dict) else None
    if slot_date and db.is_date_blocked(slot_date):
        return jsonify({"ok": False, "error": "blocked_delivery_date"}), 400

    stock_by_id = {p["id"]: p["stock"] for p in read_products_from_excel()}
    for i in items:
        pid = int(i["id"])
        requested = int(i.get("qty", 1))
        available = stock_by_id.get(pid, 0)
        if requested > available:
            return jsonify({
                "ok": False, "error": "insufficient_stock", "product_id": pid, "available": available,
            }), 409

    subtotal = sum(int(i.get("price", 0)) * int(i.get("qty", 1)) for i in items)

    discount_amount = 0
    applied_coupon = None
    if coupon_code:
        coupon = db.get_coupon(coupon_code)
        if coupon and coupon["active"]:
            discount_amount = round(subtotal * coupon["discount_percent"] / 100)
            applied_coupon = coupon["code"]
        else:
            return jsonify({"ok": False, "error": "invalid_coupon"}), 400

    is_first_order = FREE_SHIPPING_ON_FIRST_ORDER and db.count_orders_for_user(uid) == 0
    shipping_fee = 0 if is_first_order else STANDARD_SHIPPING_FEE
    total = max(0, subtotal - discount_amount) + shipping_fee

    order = {
        "id": str(uuid.uuid4())[:8],
        "user_id": uid,
        "customer": customer,
        "items": items,
        "delivery": delivery,
        "subtotal": subtotal,
        "shipping_fee": shipping_fee,
        "discount_amount": discount_amount,
        "coupon_code": applied_coupon,
        "free_shipping_applied": is_first_order,
        "total": total,
        "status": "pending_payment",
        "created_at": datetime.utcnow().isoformat(),
    }
    db.create_order(order)
    return jsonify({"ok": True, "order": order})


@app.route("/api/orders/<order_id>/pay", methods=["POST"])
def pay_order(order_id):
    """
    Placeholder payment confirmation. Wire this to a real gateway (e.g.
    Zarinpal / IDPay) by calling their "request payment" API here,
    redirecting the browser to their page, and verifying the callback
    before marking the order paid. Until real credentials are supplied,
    this simulates a successful payment so the rest of the flow (invoice +
    accounting) works.
    """
    order_row = db.get_order(order_id)
    if not order_row:
        return jsonify({"ok": False, "error": "order_not_found"}), 404
    order = db.order_to_dict(order_row)

    stock_sync_error = None
    try:
        decrement_stock(order["items"])
    except Exception as exc:
        stock_sync_error = str(exc)

    db.mark_order_paid(order_id, stock_sync_error)
    order = db.order_to_dict(db.get_order(order_id))

    invoice = {
        "id": f"INV-{db.count_invoices() + 1001}",
        "order_id": order["id"],
        "customer": order["customer"],
        "subtotal": order["subtotal"],
        "shipping_fee": order["shipping_fee"],
        "discount_amount": order["discount_amount"],
        "total": order["total"],
        "issued_at": datetime.utcnow().isoformat(),
    }
    db.create_invoice(invoice)

    return jsonify({"ok": True, "order": order, "invoice": invoice})


@app.route("/api/orders/mine", methods=["GET"])
def my_orders():
    uid = session.get("user_id")
    if not uid:
        return jsonify({"ok": False, "error": "login_required"}), 401
    return jsonify({"ok": True, "orders": db.get_orders_for_user(uid)})


@app.route("/api/admin/orders", methods=["GET"])
@require_admin
def admin_orders():
    return jsonify({"ok": True, "orders": db.get_all_orders()})


@app.route("/api/admin/orders/<order_id>/status", methods=["POST"])
@require_admin
def admin_update_order_status(order_id):
    body = request.get_json(force=True, silent=True) or {}
    status = (body.get("status") or "").strip()
    allowed = {"pending_payment", "paid", "shipped", "delivered", "cancelled"}
    if status not in allowed:
        return jsonify({"ok": False, "error": "invalid_status"}), 400
    if not db.get_order(order_id):
        return jsonify({"ok": False, "error": "order_not_found"}), 404
    db.set_order_status(order_id, status)
    return jsonify({"ok": True})


@app.route("/api/accounting/summary", methods=["GET"])
@require_admin
def accounting_summary():
    orders = db.get_all_orders()
    products = read_products_from_excel()

    paid_orders = [o for o in orders if o["status"] == "paid"]
    revenue = sum(o["total"] for o in paid_orders)
    shipping_collected = sum(o["shipping_fee"] for o in paid_orders)
    discounts_given = sum(o["discount_amount"] for o in paid_orders)
    inventory_value = sum(p["price"] * p["stock"] for p in products)
    out_of_stock = [p["name_fa"] for p in products if p["stock"] == 0]
    low_stock = [p["name_fa"] for p in products if 0 < p["stock"] <= 3]

    return jsonify({
        "ok": True,
        "revenue_toman": revenue,
        "shipping_collected_toman": shipping_collected,
        "discounts_given_toman": discounts_given,
        "paid_orders": len(paid_orders),
        "pending_orders": len([o for o in orders if o["status"] == "pending_payment"]),
        "invoices_issued": db.count_invoices(),
        "inventory_value_toman": inventory_value,
        "out_of_stock_products": out_of_stock,
        "low_stock_products": low_stock,
        "open_repairs": len([r for r in db.get_all_repairs() if r["status"] not in ("done", "cancelled")]),
    })


# --------------------------------------------------------------------------- #
# repairs
# --------------------------------------------------------------------------- #
@app.route("/api/repairs", methods=["POST"])
def create_repair_route():
    body = request.get_json(force=True, silent=True) or {}
    name = (body.get("name") or "").strip()
    phone = (body.get("phone") or "").strip()
    device_type = (body.get("device_type") or "").strip()
    issue = (body.get("issue") or "").strip()
    if not name or not phone or not device_type or not issue:
        return jsonify({"ok": False, "error": "invalid_input"}), 400

    repair = {
        "id": str(uuid.uuid4())[:8],
        "user_id": session.get("user_id"),
        "name": name, "phone": phone, "device_type": device_type, "issue": issue,
        "created_at": datetime.utcnow().isoformat(),
    }
    db.create_repair(repair)
    return jsonify({"ok": True, "repair": repair})


@app.route("/api/repairs/mine", methods=["GET"])
def my_repairs():
    uid = session.get("user_id")
    if not uid:
        return jsonify({"ok": False, "error": "login_required"}), 401
    return jsonify({"ok": True, "repairs": db.get_repairs_for_user(uid)})


@app.route("/api/admin/repairs", methods=["GET"])
@require_admin
def admin_repairs():
    return jsonify({"ok": True, "repairs": db.get_all_repairs()})


@app.route("/api/admin/repairs/<repair_id>/status", methods=["POST"])
@require_admin
def admin_update_repair_status(repair_id):
    body = request.get_json(force=True, silent=True) or {}
    status = (body.get("status") or "").strip()
    allowed = {"new", "in_progress", "done", "cancelled"}
    if status not in allowed:
        return jsonify({"ok": False, "error": "invalid_status"}), 400
    db.set_repair_status(repair_id, status)
    return jsonify({"ok": True})


# --------------------------------------------------------------------------- #
# static frontend
# --------------------------------------------------------------------------- #
@app.route("/")
def serve_index():
    return send_from_directory(FRONTEND_DIR, "index.html")


@app.route("/<path:path>")
def serve_static(path):
    return send_from_directory(FRONTEND_DIR, path)


if __name__ == "__main__":
    app.run(debug=True, port=5000)
